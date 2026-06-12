import io
import threading
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, BackgroundTasks, Depends, Query
from fastapi.responses import StreamingResponse

from backend.database.connection import get_connection, init_db
from backend.middleware.auth import require_auth
from backend.database.repositories import prices as price_repo
from backend.ai.price_extractor import (
    extract_prices_from_all_imap,
    extract_prices_from_email,
)
from backend.config import settings

router = APIRouter(prefix="/api/prices", tags=["prices"], dependencies=[Depends(require_auth)])

# In-memory job state (single-process app)
_job: dict = {"running": False, "processed": 0, "skipped": 0, "total_quotes": 0,
               "errors": 0, "folders_scanned": [], "done": False, "error": None}
_job_lock = threading.Lock()


def _run_extraction():
    """Runs in a background thread. Opens its own DB connection."""
    with _job_lock:
        _job.update(running=True, processed=0, skipped=0, total_quotes=0,
                    errors=0, folders_scanned=[], done=False, error=None)
    try:
        init_db(settings.database_path)
        conn = get_connection()
        result = extract_prices_from_all_imap(conn, progress_cb=_progress_cb)
        conn.close()
        with _job_lock:
            _job.update(**result, running=False, done=True)
    except Exception as e:
        with _job_lock:
            _job.update(running=False, done=True, error=str(e))


def _progress_cb(processed: int, skipped: int, total_quotes: int,
                  errors: int, folder: str):
    with _job_lock:
        _job["processed"] = processed
        _job["skipped"] = skipped
        _job["total_quotes"] = total_quotes
        _job["errors"] = errors
        if folder and folder not in _job["folders_scanned"]:
            _job["folders_scanned"].append(folder)


@router.get("")
def list_prices(
    category: str | None = Query(None),
    vendor: str | None = Query(None),
    search: str | None = Query(None),
    conn=Depends(get_connection),
):
    rows = price_repo.list_all(conn, category=category, vendor=vendor, search=search)
    categories = price_repo.list_categories(conn)
    return {
        "items": [dict(r) for r in rows],
        "categories": categories,
        "total": len(rows),
    }


@router.post("/extract")
def extract_all():
    """Start price extraction as a background job. Returns immediately."""
    with _job_lock:
        if _job["running"]:
            return {"started": False, "message": "Already running"}
    thread = threading.Thread(target=_run_extraction, daemon=True)
    thread.start()
    return {"started": True, "message": "Extraction started"}


@router.get("/extract/status")
def extract_status():
    with _job_lock:
        return dict(_job)


@router.post("/extract/{email_id}")
def extract_one(email_id: int, conn=Depends(get_connection)):
    from backend.database.repositories import prices as pr
    pr.delete_by_email(conn, email_id)
    quotes = extract_prices_from_email(conn, email_id)
    return {"email_id": email_id, "quotes_found": len(quotes), "quotes": quotes}


@router.get("/export")
def export_excel(
    category: str | None = Query(None),
    vendor: str | None = Query(None),
    search: str | None = Query(None),
    conn=Depends(get_connection),
):
    rows = price_repo.list_all(conn, category=category, vendor=vendor, search=search)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price List"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = [
        "Category", "Product Name", "Part Number", "Unit Price", "Currency",
        "Quantity", "Vendor", "Vendor Email", "Validity Date", "Notes",
        "Source Email", "Extracted At",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    col_widths = [14, 40, 20, 12, 10, 10, 22, 30, 14, 40, 35, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    cat_colors = {
        "Cisco":      "E8F4FD",
        "Fortinet":   "FEF3C7",
        "HP":         "F0FDF4",
        "Microsoft":  "EDE9FE",
        "Lenovo":     "FFF7ED",
        "Dell":       "FDF2F8",
        "Aruba":      "ECFDF5",
        "Palo Alto":  "FEF2F2",
        "Other":      "F8FAFC",
    }

    for row_num, r in enumerate(rows, 2):
        values = [
            r["category"] or "Other",
            r["product_name"],
            r["part_number"] or "",
            r["unit_price"],
            r["currency"] or "USD",
            r["quantity"] or "",
            r["vendor_name"] or "",
            r["vendor_email"] or "",
            r["validity_date"] or "",
            r["notes"] or "",
            r["effective_subject"] or "",
            r["extracted_at"],
        ]
        cat = r["category"] or "Other"
        fill_color = cat_colors.get(cat, "F8FAFC")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = row_fill
            if col == 4 and val is not None:
                cell.number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.cell(row=1, column=1, value="AAKE Price List Summary").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=3, column=1, value=f"Total items: {len(rows)}")
    ws2.cell(row=5, column=1, value="Category").font = Font(bold=True)
    ws2.cell(row=5, column=2, value="Items").font = Font(bold=True)

    cat_counts: dict[str, int] = {}
    for r in rows:
        cat = r["category"] or "Other"
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    for i, (cat, count) in enumerate(sorted(cat_counts.items()), 6):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=count)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"AAKE_Price_List_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
